#!/usr/bin/env python3
"""Validate GTM audit completion-gate reconciliation rows.

Input may be:
- CSV exported from the Workstream Reconciliation tab.
- JSON list of row objects, or {"rows": [...]}.
- XLSX workbook with a sheet named "18b Workstream Reconciliation" or any
  sheet whose name contains "Reconciliation".

Use --strict-evidence with XLSX workbooks that claim full audit or cleanup-plan
completion. Strict mode also checks Semantic Object Matrix and Custom Code
Semantic Review evidence and rejects cleanup-operation placeholders that defer
audit work into execution.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import zipfile
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple
from xml.etree import ElementTree

from gtm_workbook import (
    column_index,
    find_sheet,
    load_xlsx_workbook,
    normalize_header,
    normalize_sheet_name,
    rows_from_values,
    workbook_target_path,
    xml_text,
)


COUNT_FIELDS = [
    "total_source_count",
    "inventoried_count",
    "dependency_mapped_count",
    "measurement_diagnosed_count",
    "semantically_validated_count",
    "cleanup_decision_count",
    "deferred_count",
    "not_applicable_count",
    "user_excluded_count",
    "unresolved_count",
]

REQUIRED_FIELDS = ["workstream", "object_family"] + COUNT_FIELDS

SEMANTIC_MATRIX_REQUIRED_FIELDS = [
    "object_id",
    "object_name",
    "layer",
    "vendor_or_family",
    "inferred_business_role",
    "decision_outcome",
    "conversion_hierarchy",
    "platform_role",
    "expected_data_contract",
    "depth_required",
    "depth_completed",
    "trigger_context_status",
    "configuration_logic_status",
    "source_or_code_logic_status",
    "consent_or_server_status",
    "evidence_level",
    "semantic_status",
    "confidence",
    "runtime_qa_required",
    "blocker_or_next_evidence",
]

CUSTOM_CODE_REQUIRED_FIELDS = [
    "layer",
    "object_id",
    "object_name",
    "type",
    "role_category",
    "purpose",
    "export_review_completed",
    "trigger_or_consumer_context",
    "consent_assumption",
    "external_urls_storage_cookie_dom_datalayer_side_effects",
    "variable_references",
    "expected_output_or_side_effect",
    "runtime_risks",
    "semantic_status",
    "cleanup_recommendation",
    "qa_method",
    "blocker",
]

D3_REQUIRED_FIELDS = [
    "d3_inputs_or_sources",
    "d3_logic_summary",
    "d3_output_or_side_effect",
    "d3_consumer_expectation",
    "d3_correctness_decision",
]

SEMANTIC_SUMMARY_FIELDS = [
    "configuration_logic_status",
    "source_or_code_logic_status",
    "d3_logic_summary",
    "d3_output_or_side_effect",
    "d3_correctness_decision",
]

CUSTOM_CODE_SUMMARY_FIELDS = [
    "purpose",
    "external_urls_storage_cookie_dom_datalayer_side_effects",
    "expected_output_or_side_effect",
    "runtime_risks",
    "cleanup_recommendation",
]

PLACEHOLDER_PATTERNS = [
    re.compile(r"\bperform\s+line[- ]level(?:\s+custom[- ]code)?\s+review\b", re.I),
    re.compile(r"\breview\s+custom\s+code\b", re.I),
    re.compile(r"\bcheck\s+(?:the\s+)?variables?\b", re.I),
    re.compile(r"\bvalidate\s+trigger\s+logic\b", re.I),
]

INCOMPLETE_DEPTH_PATTERNS = [
    re.compile(r"\bd3\s*/\s*d4\s+blocked\b", re.I),
    re.compile(r"\bd3\s+(?:required|needed)\b", re.I),
    re.compile(r"\bstatic\s+scan\s+only\b", re.I),
    re.compile(r"\b(?:full\s+)?code\s+walkthrough\s+(?:required|needed)\b", re.I),
    re.compile(r"\breview\s+later\b", re.I),
]

GENERIC_SUMMARY_PATTERNS = [
    re.compile(r"\bcustom\s+code\s+inspected\b", re.I),
    re.compile(r"\bconfiguration\s+reviewed\b", re.I),
    re.compile(r"\bcode\s+scanned\b", re.I),
    re.compile(r"\bexternal\s+url\s+found\b", re.I),
    re.compile(r"\bdatalayer\s+push\s+detected\b", re.I),
    re.compile(r"\bno\s+obvious\s+browser\s+side\s+effect\b", re.I),
    re.compile(r"\bno\s+issue\s+found\b", re.I),
    re.compile(r"\bsee\s+(?:the\s+)?config\b", re.I),
    re.compile(r"\bsee\s+(?:the\s+)?export\b", re.I),
    re.compile(r"\bstatic\s+scan\s+completed\b", re.I),
    re.compile(r"\breviewed\s+manually\b", re.I),
]

GENERIC_D3_VALUES = {
    "",
    "n/a",
    "na",
    "not applicable",
    "see export",
    "see config",
    "static scan",
    "static scan completed",
    "runtime required",
    "runtime qa required",
    "more info needed",
    "blocked",
    "pending",
}


def to_int(row: Dict[str, Any], field: str) -> Tuple[int, str | None]:
    raw = row.get(field, "")
    if raw is None or raw == "":
        return 0, f"missing count '{field}'"
    try:
        return int(float(str(raw).strip())), None
    except ValueError:
        return 0, f"invalid count '{field}'={raw!r}"


def load_csv(path: Path) -> List[Dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        rows = []
        for raw in reader:
            rows.append({normalize_header(k): v for k, v in raw.items()})
        return rows


def load_json(path: Path) -> List[Dict[str, Any]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(data, dict) and isinstance(data.get("rows"), list):
        data = data["rows"]
    if not isinstance(data, list):
        raise ValueError("JSON input must be a list of rows or {'rows': [...]}")
    return [{normalize_header(k): v for k, v in row.items()} for row in data]


def load_xlsx_stdlib(path: Path) -> List[Dict[str, Any]]:
    ns = {
        "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "pkgrel": "http://schemas.openxmlformats.org/package/2006/relationships",
    }

    with zipfile.ZipFile(path) as archive:
        workbook = ElementTree.fromstring(archive.read("xl/workbook.xml"))
        rels = ElementTree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))

        rel_targets = {
            rel.attrib["Id"]: rel.attrib["Target"]
            for rel in rels.findall("pkgrel:Relationship", ns)
        }

        sheet_target = None
        for sheet in workbook.findall("main:sheets/main:sheet", ns):
            name = sheet.attrib.get("name", "")
            rel_id = sheet.attrib.get(f"{{{ns['rel']}}}id")
            if name == "18b Workstream Reconciliation" or "reconciliation" in name.lower():
                target = rel_targets.get(rel_id or "")
                if not target:
                    continue
                sheet_target = workbook_target_path(target)
                break
        if not sheet_target:
            raise ValueError("No reconciliation sheet found")

        shared_strings: List[str] = []
        if "xl/sharedStrings.xml" in archive.namelist():
            shared_root = ElementTree.fromstring(archive.read("xl/sharedStrings.xml"))
            for item in shared_root.findall("main:si", ns):
                shared_strings.append(xml_text(item))

        sheet_root = ElementTree.fromstring(archive.read(sheet_target))
        parsed_rows: List[List[str]] = []
        for row in sheet_root.findall("main:sheetData/main:row", ns):
            values: List[str] = []
            for cell in row.findall("main:c", ns):
                ref = cell.attrib.get("r", "")
                index = column_index(ref) if ref else len(values)
                while len(values) <= index:
                    values.append("")
                cell_type = cell.attrib.get("t", "")
                if cell_type == "s":
                    raw = xml_text(cell.find("main:v", ns))
                    values[index] = shared_strings[int(raw)] if raw else ""
                elif cell_type == "inlineStr":
                    values[index] = xml_text(cell.find("main:is", ns))
                else:
                    values[index] = xml_text(cell.find("main:v", ns))
            if any(value != "" for value in values):
                parsed_rows.append(values)

    if not parsed_rows:
        return []
    headers = [normalize_header(value) for value in parsed_rows[0]]
    rows = []
    for values in parsed_rows[1:]:
        rows.append({headers[i]: values[i] for i in range(min(len(headers), len(values)))})
    return rows


def load_xlsx_openpyxl(path: Path) -> List[Dict[str, Any]]:
    try:
        import openpyxl  # type: ignore
    except ImportError as exc:
        raise RuntimeError("openpyxl is unavailable") from exc

    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet_name = None
    if "18b Workstream Reconciliation" in workbook.sheetnames:
        sheet_name = "18b Workstream Reconciliation"
    else:
        for candidate in workbook.sheetnames:
            if "reconciliation" in candidate.lower():
                sheet_name = candidate
                break
    if not sheet_name:
        raise ValueError("No reconciliation sheet found")

    sheet = workbook[sheet_name]
    rows_iter = sheet.iter_rows(values_only=True)
    headers = [normalize_header(v) for v in next(rows_iter, [])]
    rows = []
    for values in rows_iter:
        if not values or all(v in (None, "") for v in values):
            continue
        rows.append({headers[i]: values[i] for i in range(min(len(headers), len(values)))})
    return rows


def load_xlsx(path: Path) -> List[Dict[str, Any]]:
    try:
        return load_xlsx_openpyxl(path)
    except Exception as first_exc:  # noqa: BLE001 - fallback should cover missing optional deps.
        try:
            return load_xlsx_stdlib(path)
        except Exception as second_exc:  # noqa: BLE001 - preserve both failure reasons.
            raise RuntimeError(
                f"Unable to read XLSX with openpyxl or stdlib fallback. "
                f"openpyxl error: {first_exc}; fallback error: {second_exc}"
            ) from second_exc


def load_rows(path: Path) -> List[Dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return load_csv(path)
    if suffix == ".json":
        return load_json(path)
    if suffix == ".xlsx":
        return load_xlsx(path)
    raise ValueError("Unsupported file type. Use .csv, .json, or .xlsx")


def validate_rows(rows: Iterable[Dict[str, Any]]) -> Tuple[List[str], List[str]]:
    errors: List[str] = []
    warnings: List[str] = []
    row_list = list(rows)
    if not row_list:
        return ["no reconciliation rows found"], warnings

    for index, row in enumerate(row_list, start=2):
        label = f"row {index}"
        missing = [field for field in REQUIRED_FIELDS if field not in row]
        if missing:
            errors.append(f"{label}: missing required fields: {', '.join(missing)}")
            continue

        counts: Dict[str, int] = {}
        for field in COUNT_FIELDS:
            value, problem = to_int(row, field)
            counts[field] = value
            if problem:
                errors.append(f"{label}: {problem}")

        semantic_total = (
            counts["semantically_validated_count"]
            + counts["deferred_count"]
            + counts["not_applicable_count"]
            + counts["user_excluded_count"]
        )
        if counts["total_source_count"] != semantic_total:
            errors.append(
                f"{label}: semantic coverage mismatch: total_source_count="
                f"{counts['total_source_count']} but semantically_validated + deferred "
                f"+ not_applicable + user_excluded = {semantic_total}"
            )
        if counts["unresolved_count"] != 0:
            errors.append(f"{label}: unresolved_count is {counts['unresolved_count']}")
        if counts["inventoried_count"] < counts["total_source_count"]:
            warnings.append(f"{label}: inventoried_count is below total_source_count")
        if counts["dependency_mapped_count"] < counts["semantically_validated_count"]:
            warnings.append(f"{label}: dependency_mapped_count is below semantically_validated_count")
        if counts["measurement_diagnosed_count"] < counts["semantically_validated_count"]:
            errors.append(
                f"{label}: measurement_diagnosed_count is below semantically_validated_count"
            )
        if counts["cleanup_decision_count"] < counts["semantically_validated_count"]:
            warnings.append(f"{label}: cleanup_decision_count is below semantically_validated_count")

    return errors, warnings


def field_is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def validate_required_table(
    rows: List[Dict[str, Any]], required_fields: List[str], label: str
) -> List[str]:
    errors: List[str] = []
    if not rows:
        return [f"{label}: no data rows found"]

    available = set(rows[0])
    missing = [field for field in required_fields if field not in available]
    if missing:
        errors.append(f"{label}: missing required fields: {', '.join(missing)}")
        return errors

    key_fields = [
        field
        for field in ("object_id", "object_name", "layer", "semantic_status")
        if field in required_fields
    ]
    key_fields.extend(
        field
        for field in (
            "inferred_business_role",
            "decision_outcome",
            "conversion_hierarchy",
            "platform_role",
            "expected_data_contract",
        )
        if field in required_fields
    )
    for index, row in enumerate(rows, start=2):
        for field in key_fields:
            if field_is_blank(row.get(field)):
                errors.append(f"{label} row {index}: blank {field}")
    return errors


def depth_tokens(value: Any) -> set[str]:
    return {match.group(0).upper() for match in re.finditer(r"\bD[1-4]\b", str(value or ""), re.I)}


def generic_or_blank(value: Any) -> bool:
    text = str(value or "").strip().lower()
    return text in GENERIC_D3_VALUES


def validate_semantic_depth_rows(rows: List[Dict[str, Any]], label: str) -> List[str]:
    errors: List[str] = []
    available = set(rows[0]) if rows else set()
    missing_d3_columns = [field for field in D3_REQUIRED_FIELDS if field not in available]

    for index, row in enumerate(rows, start=2):
        required = depth_tokens(row.get("depth_required"))
        completed = depth_tokens(row.get("depth_completed"))
        missing_depths = sorted(required.intersection({"D1", "D2", "D3"}) - completed)
        if missing_depths:
            errors.append(
                f"{label} row {index}: depth_required includes "
                f"{', '.join(missing_depths)} but depth_completed does not"
            )

        depth_completed_text = str(row.get("depth_completed") or "")
        source_logic_text = str(row.get("source_or_code_logic_status") or "")
        for pattern in INCOMPLETE_DEPTH_PATTERNS:
            if pattern.search(depth_completed_text) or pattern.search(source_logic_text):
                errors.append(
                    f"{label} row {index}: incomplete-depth wording found: {pattern.pattern!r}"
                )

        if "D3" in required:
            if missing_d3_columns:
                errors.append(
                    f"{label} row {index}: D3 required but matrix is missing columns: "
                    f"{', '.join(missing_d3_columns)}"
                )
                continue
            if "D3" not in completed:
                continue
            for field in D3_REQUIRED_FIELDS:
                if generic_or_blank(row.get(field)):
                    errors.append(f"{label} row {index}: D3 field {field} is blank or generic")
    return errors


def validate_summary_quality(
    rows: List[Dict[str, Any]], fields: List[str], label: str
) -> List[str]:
    errors: List[str] = []
    if not rows:
        return errors

    available = set(rows[0])
    target_fields = [field for field in fields if field in available]
    for index, row in enumerate(rows, start=2):
        for field in target_fields:
            text = str(row.get(field) or "")
            for pattern in GENERIC_SUMMARY_PATTERNS:
                if pattern.search(text):
                    errors.append(
                        f"{label} row {index} field {field}: generic summary phrase "
                        f"{pattern.pattern!r}; explain category, source/input, "
                        "logic/action, output or side effect, and judgment"
                    )
    return errors


def custom_code_required(reconciliation_rows: Iterable[Dict[str, Any]]) -> bool:
    for row in reconciliation_rows:
        label = f"{row.get('workstream', '')} {row.get('object_family', '')}".lower()
        if "custom" not in label:
            continue
        total, problem = to_int(row, "total_source_count")
        if problem:
            continue
        not_applicable, _ = to_int(row, "not_applicable_count")
        user_excluded, _ = to_int(row, "user_excluded_count")
        if total > not_applicable + user_excluded:
            return True
    return False


def validate_custom_code_rows(rows: List[Dict[str, Any]], label: str) -> List[str]:
    errors = validate_required_table(rows, CUSTOM_CODE_REQUIRED_FIELDS, label)
    if errors:
        return errors

    for index, row in enumerate(rows, start=2):
        review_status = str(row.get("export_review_completed", "")).strip().lower()
        semantic_status = str(row.get("semantic_status", "")).strip().lower()
        if review_status not in {"yes", "not applicable", "n/a"}:
            errors.append(
                f"{label} row {index}: export_review_completed must be Yes or Not applicable"
            )
        if semantic_status in {"", "review later", "pending", "pending review"}:
            errors.append(f"{label} row {index}: semantic_status is not a decision")
    errors.extend(validate_summary_quality(rows, CUSTOM_CODE_SUMMARY_FIELDS, label))
    return errors


def validate_placeholder_language(
    workbook_rows: Dict[str, List[Dict[str, Any]]]
) -> List[str]:
    errors: List[str] = []
    target_sheet_terms = ("finding", "operation", "roadmap", "cleanup", "plan")
    for sheet_name, rows in workbook_rows.items():
        normalized = normalize_sheet_name(sheet_name)
        if not any(term in normalized for term in target_sheet_terms):
            continue
        for row_index, row in enumerate(rows, start=2):
            for field, value in row.items():
                text = str(value or "")
                for pattern in PLACEHOLDER_PATTERNS:
                    if pattern.search(text):
                        errors.append(
                            f"{sheet_name} row {row_index} field {field}: "
                            f"deferred audit-work placeholder {pattern.pattern!r}"
                        )
    return errors


def validate_strict_evidence(path: Path, reconciliation_rows: List[Dict[str, Any]]) -> Tuple[List[str], List[str]]:
    errors: List[str] = []
    warnings: List[str] = []
    if path.suffix.lower() != ".xlsx":
        return ["--strict-evidence requires an XLSX workbook"], warnings

    workbook_rows = load_xlsx_workbook(path)

    semantic_name, semantic_rows = find_sheet(workbook_rows, ["semantic", "matrix"])
    if not semantic_name:
        errors.append("strict evidence: missing Semantic Object Matrix sheet")
    else:
        errors.extend(
            validate_required_table(
                semantic_rows,
                SEMANTIC_MATRIX_REQUIRED_FIELDS,
                f"{semantic_name}",
            )
        )
        if semantic_rows:
            errors.extend(validate_semantic_depth_rows(semantic_rows, f"{semantic_name}"))
            errors.extend(
                validate_summary_quality(
                    semantic_rows,
                    SEMANTIC_SUMMARY_FIELDS,
                    f"{semantic_name}",
                )
            )

    custom_name, custom_rows = find_sheet(workbook_rows, ["custom", "code"])
    if custom_code_required(reconciliation_rows) and not custom_name:
        errors.append(
            "strict evidence: reconciliation indicates custom-code scope, "
            "but no Custom Code Semantic Review sheet was found"
        )
    elif custom_name:
        errors.extend(validate_custom_code_rows(custom_rows, f"{custom_name}"))

    errors.extend(validate_placeholder_language(workbook_rows))
    if not custom_code_required(reconciliation_rows) and not custom_name:
        warnings.append("strict evidence: no custom-code review sheet found; treated as not in scope")

    return errors, warnings


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input", type=Path, help="CSV, JSON, or XLSX reconciliation file")
    parser.add_argument(
        "--strict-evidence",
        action="store_true",
        help=(
            "For full audit/cleanup-plan XLSX workbooks, also validate Semantic "
            "Object Matrix and Custom Code Semantic Review evidence."
        ),
    )
    args = parser.parse_args()

    try:
        rows = load_rows(args.input)
        errors, warnings = validate_rows(rows)
        if args.strict_evidence:
            strict_errors, strict_warnings = validate_strict_evidence(args.input, rows)
            errors.extend(strict_errors)
            warnings.extend(strict_warnings)
    except Exception as exc:  # noqa: BLE001 - CLI should report any loading problem.
        print(f"ERROR: {exc}", file=sys.stderr)
        return 2

    for warning in warnings:
        print(f"WARNING: {warning}")
    if errors:
        for error in errors:
            print(f"ERROR: {error}", file=sys.stderr)
        print(f"Gate status: FAIL ({len(errors)} error(s), {len(warnings)} warning(s))")
        return 1

    print(f"Gate status: PASS ({len(rows)} row(s), {len(warnings)} warning(s))")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
